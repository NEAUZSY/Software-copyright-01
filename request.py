import os
import time
from datetime import datetime

import requests
import openpyxl as xl
from lxml import etree
from utils.header import get_headers

# 选择爬取的页数 默认为20
PAGES = 20
# 是否在本地保存html文件
SAVE_HTML = False

BASE_URL = 'http://q.10jqka.com.cn/index/index/board/all/field/zdf/order/desc/page/{}/'


def get(base_url):
    # 获取请求头
    header = get_headers()
    # 判断本地是否已经存在之前爬取的网页
    current_time = datetime.now().strftime('%Y-%m-%d-%H时%M分%S秒')
    html = ''
    # 保存爬取的源网页文件
    if SAVE_HTML:
        html_name = 'stock-{}.html'.format(current_time)

        if os.path.exists(html_name):
            # 如果已经存在之前爬取的网页便将其赋值给html
            with open(html_name, encoding='GBK') as f:
                html = f.read()
        else:
            # 否则在网络上进行爬取，并存储在html变量中
            for i in range(1, PAGES + 1):
                # 获取真正的url
                url = base_url.format(i)

                # 进行请求
                resp = requests.get(url, headers=header)
                # print(resp.status_cod)
                # assert resp.status_code == 200

                # 获取请求内容
                html += resp.text

                # 暂停一会反爬
                print('完成了{}页抓取'.format(i))
                time.sleep(0.1)

            with open(html_name, 'w', encoding='GBK') as f:
                f.write(html)
        # 返回网页文件
    # 不保存html文件 爬取后直接解析并存储到excel中
    else:
        for i in range(1, PAGES + 1):
            # 获取真正的url
            url = base_url.format(i)

            # 进行请求
            resp = requests.get(url, headers=header)
            # print(resp.status_cod)
            # assert resp.status_code == 200

            # 获取请求内容
            html += resp.text

            # 暂停一会反爬
            print('完成了{}页抓取'.format(i))
            time.sleep(0.1)
    return html


def parse(html):
    # 创建根节点
    root = etree.HTML(html)
    # 找到今日个股行情信息
    tbody = root.xpath('//table/tbody/tr')
    data = []
    for tr in tbody:
        # 初始化每一行的列表
        row = [str(tr.xpath('./td[1]/text()')[0]),
               str(tr.xpath('./td[2]/a/text()')[0]),
               str(tr.xpath('./td[3]/a/text()')[0])]
        for i in range(4, 15):
            target = r'./td[{}]/text()'.format(i)
            row.append(str(tr.xpath(target)[0]))
        data.append(row)
    # print(data)
    return data


def save_data(data):
    # 获取当前日期
    current_time = datetime.now().strftime('%Y-%m-%d-%H时%M分%S秒')
    # date = current_time[:10]

    book_name = '股价.xlsx'

    # 判断本地是否已经存在文件
    if os.path.exists(book_name):
        # 如果已经存在之前爬取的网页便将其赋值给html
        wb = xl.load_workbook(filename=book_name)
        ws = wb.create_sheet(title=current_time)
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.title = current_time

    title = ['序号', '代码', '名称', '现价', '涨跌幅（%）', '涨跌', '涨速（%）',
             '换手（%）', '量比', '振幅（%）', '成交额', '流通股', '流通市值', '市盈率']
    ws.append(title)
    for row in data:
        ws.append(row)
    wb.save(book_name)
    print('保存成功')


def main():
    html = get(BASE_URL)
    data = parse(html)
    save_data(data)


if __name__ == '__main__':
    main()
