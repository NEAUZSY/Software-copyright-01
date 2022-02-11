import os
import time
from datetime import datetime

import openpyxl as xl
from lxml import etree
from requests import get
from utils.header import get_headers
from threading import Thread

from icon import img
from utils.log import log
from utils import global_value as gl

log('下载程序包引入完成')
# 选择爬取的页数 默认为20
PAGES = 10
# 是否在本地保存html文件
SAVE_HTML = False

BASE_URL = 'http://q.10jqka.com.cn/index/index/board/all/field/zdf/order/desc/page/{}/'

log('全局变量及包声明完成')


def get_self(base_url, pages, is_save_html):
    # 获取请求头
    log('获取请求头')
    header = get_headers()
    # 判断本地是否已经存在之前爬取的网页
    log('判断本地是否已经存在之前爬取的网页')
    current_time = datetime.now().strftime('%Y-%m-%d-%H时%M分%S秒')
    html = ''
    # 保存爬取的源网页文件
    # print(is_save_html)
    if is_save_html:
        log('保存原网页')
        html_name = 'stock-{}.html'.format(current_time)

        if os.path.exists(html_name):
            # 如果已经存在之前爬取的网页便将其赋值给html
            with open(html_name, encoding='GBK') as f:
                html = f.read()
        else:
            # 否则在网络上进行爬取，并存储在html变量中
            for i in range(1, pages + 1):
                # 获取真正的url
                url = base_url.format(i)

                # 进行请求
                resp = get(url, headers=header)
                while not resp.status_code == 200:
                    log('请求失败,正在重新请求')
                    time.sleep(0.5)
                    resp = get(url, headers=header)
                    pass

                # 获取请求内容
                html += resp.text
                log('完成了{}页抓取'.format(i))

                # 暂停一会反爬
                gl.lock.acquire()
                gl.set_value('page', i)
                gl.lock.release()
                time.sleep(0.1)

            with open(html_name, 'w', encoding='GBK') as f:
                f.write(html)
        # 返回网页文件
    # 不保存html文件 爬取后直接解析并存储到excel中
    else:
        for i in range(1, pages + 1):
            # 获取真正的url
            url = base_url.format(i)

            # 进行请求
            resp = get(url, headers=header)
            # print(resp.status_cod)
            # assert resp.status_code == 200

            # 获取请求内容
            html += resp.text

            # 暂停一会反爬
            log('完成了{}页抓取'.format(i))
            gl.lock.acquire()
            gl.set_value('page', i)
            gl.lock.release()
            time.sleep(0.1)
    log('完成所有网页爬取')
    return html


def parse(html):
    log('开始解析')
    # 创建根节点
    root_node = etree.HTML(html)
    # 找到今日个股行情信息
    tbody = root_node.xpath('//table/tbody/tr')
    data = []
    for tr in tbody:
        # 初始化每一行的列表
        row = [str(tr.xpath('./td[1]/text()')[0]),
               str(tr.xpath('./td[2]/a/text()')[0]),
               str(tr.xpath('./td[3]/a/text()')[0])]
        for i in range(4, 15):
            target = r'./td[{}]/text()'.format(i)
            row.append(str(tr.xpath(target)[0]))
        data.append(row)
    # print(data)
    log('解析完成')
    return data


def save_data(data):
    # 获取当前日期
    log('储存信息到xlsx文件')
    current_time = datetime.now().strftime('%Y-%m-%d-%H时%M分%S秒')
    # date = current_time[:10]

    book_name = '股价.xlsx'

    # 判断本地是否已经存在文件
    if os.path.exists(book_name):
        # 如果已经存在之前爬取的网页便将其赋值给html
        log('如果已经存在之前爬取的网页便将其赋值给html')
        wb = xl.load_workbook(filename=book_name)
        ws = wb.create_sheet(title=current_time)
    else:
        log('本地没有数据文件则新建')
        wb = xl.Workbook()
        ws = wb.active
        ws.title = current_time

    title = ['序号', '代码', '名称', '现价', '涨跌幅（%）', '涨跌', '涨速（%）',
             '换手（%）', '量比', '振幅（%）', '成交额', '流通股', '流通市值', '市盈率']
    ws.append(title)
    log('开始保存文件')
    for row in data:
        ws.append(row)
    wb.save(book_name)
    log('保存完成')
    # print('保存成功')


def master(is_save_html, pages):
    html = get_self(BASE_URL, pages, is_save_html)
    data = parse(html)
    save_data(data)


def ppp():
    while True:
        print(gl.get_value('page'))
        time.sleep(0.2)


if __name__ == '__main__':
    gl._init()
    t1 = Thread(target=master, args=(SAVE_HTML, PAGES,))
    t2 = Thread(target=ppp)
    t1.start()
    t2.start()
